import openpyxl

input_file = "DATA.xlsx"
wb = openpyxl.load_workbook(input_file)
data_sheet = wb['DATA']
result_sheet = wb['RESULT']

max_row = data_sheet.max_row
max_col = data_sheet.max_column

for col in range (1,max_col):
    if(data_sheet.cell(1,col).value == "NAME"):
        print ("hello")
        for row in range (2,max_row+1):
            index = row -1
            row_index = index/4 + 2
            #name1,tittle1
            if (index % 4 == 1):
                result_sheet.cell(row_index,1).value = data_sheet.cell(row,col).value
                result_sheet.cell(row_index,2).value = data_sheet.cell(row,col+1).value
            #name2,tittle2
            if (index % 4 == 2):
                result_sheet.cell(row_index,3).value = data_sheet.cell(row,col).value
                result_sheet.cell(row_index,4).value = data_sheet.cell(row,col+1).value
            #name3,tittle3
            if (index % 4 == 3):
                result_sheet.cell(row_index,5).value = data_sheet.cell(row,col).value
                result_sheet.cell(row_index,6).value = data_sheet.cell(row,col+1).value
            #name4,tittle4
            if (index % 4 == 0):
                result_sheet.cell(row_index-1,7).value = data_sheet.cell(row,col).value
                result_sheet.cell(row_index-1,8).value = data_sheet.cell(row,col+1).value
    
wb.save('updatedProduceSales.xlsx')